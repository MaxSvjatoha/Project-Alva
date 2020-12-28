
# This is the 1.1 version of the Alva algorithm, previously known as Project A

# --------------------- Section 1, library imports ---------------------------
from openpyxl import Workbook
from openpyxl import load_workbook
from difflib import SequenceMatcher
from random import choice
import string
# import time  # Useful import when debugging, e.g. time.sleep() and time.time()
import sys   # Useful import when debugging, e.g. using sys.exit()

# -------------------- Section 2, global variables ---------------------------
# Initialize all global constants, variables and lists here.
# Turn these into local variables whenever possible, but no such optimizations
# are a priority as of Alva 1.1

wb_11 = []  # The meta-list for storing 1.1 rows. Each item in the list corresponds to a row in 1.1.
wb_13 = []  # The meta-list for storing 1.3 rows. Each item in the list corresponds to a row in 1.3.
all_names = []  # The longer list of names for when both 1.1 and 1.3 are matched and are in the same row.

amount_sum = 0  # Variable to store the sum of all monetary values in a wb item.

# List used for partial matching filtering.
# The filter removes the symbols in the list from words and makes all letters lowercase.
filterlist = [" ", ",", ".", "-", "'", "&", ":", "*", "(", ")", "#"]

input_list = []                             # The initial input list for the matching functions.
matchlist = []                              # List for storing match results.
output_list = []                            # The final list with matches to be printed as the Alva output
output_row = []                             # A single row within output_list
# inputfilename_11 = "1.1 databas Alva.xlsx"     # 1.1 input file name. Uncomment if you want to preset it here.
# inputfilename_13 = "1.3 databas Alva.xlsx"     # 1.3 input file name. Uncomment if you want to preset it here.
outputfilename = "output"                   # Name of file containing the output.
orig_filename = outputfilename              # Stores a copy of the original filename.
filenamenumber = 1                          # output format is "outputfilename_filenamenumber.xlsx"
tries = 0                                   # A variable used to prevent infinite loops when printing.

# Experimental functions toggle:
prefilter = False                           # Turn the pre-filter on or off.
prefilter_erufonly = False                  # Filter away all non-ERUF rows.
postfilter = False                          # Turn the post-filter on of off.

# Matching booleans
match = True                                # Turn matching code on or off.
successful = False                          # Program status. True once printed successfully.
# Set successful boolean to "False" to write results to excel file.

# ------------------- Section 3, function definitions ------------------------

# Name: CellMake
# Makes a list of excel cells depending on input. (1, 2, 5, 6) returns ('A5', 'A6', 'B5', 'B6') for example.
# Used in ExcelWrite.

def CellMake(minlet, maxlet, minnum, maxnum):
    names = []
    letters = []
    if maxlet <= 26:
        letters = list(string.ascii_uppercase[minlet-1:maxlet])
    else:
        for i in range(minlet-1, 26):
            letters.append(string.ascii_uppercase[i])
        for i in range(26, maxlet):
            letters.append("A" + string.ascii_uppercase[i-26])
    for let in range(len(letters)):
        for num in range(minnum, maxnum+1):
            names.append(letters[let] + str(num+1))

    return names

# Name: ExcelWrite
# Takes a list of values and outputs it in an excel file using several other functions. Might be redundant.

def ExcelWrite(minlet, maxlet, minnum, maxnum, vals, name):
    minlet = minlet
    maxlet = maxlet
    minnum = minnum-1
    maxnum = maxnum-1
    vals = vals
    workbook = Workbook()

    for i in range(len(vals[0])):
        values = Unpack(vals, i)
        cell_names = CellMake(minlet+i, minlet+i, minnum, maxnum)
        print("Cell names:", cell_names)
        cell_values = values
        Write(workbook, cell_names, cell_values, name)

    successful = True

    return successful


# Name: Extract
# Extracts items from workbooks and puts them in a list. Returns said list.

def Extract(workbook, min_row, max_row, min_col, max_col, print_info):
    list = []
    if print_info:
        print("Extracting...")
    for row in workbook.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True):
       for item in row:
        list.append(item)
    return list

# Name: Filter
# Returns filtered text. Filter can contain any string and any instance of that string will be removed from the text.
# Optional mode that forces all letters to become lowercase included (true by default).

def Filter(txt, filterlist, lower=True):
    for item in filterlist:
        x = txt.split(item)
        txt = ""
        for thing in x:
            txt += thing
    txt = ""
    for item in x:
        txt += item
    if lower:
        txt = txt.lower()
    return txt

def Input_Name():
    while True:
        print("Alva running. Select one of the following commands and press enter.")
        print("1) Enter a filename for each of the files being compared")
        print("2) Use the default '1.1 databas Alva.xlsx' and '1.3 data Alva.xlsx'")
        print("0) Exit program")
        user_input = str(input("Enter input: "))
        if user_input == "1":
            inputfilename_11 = str(input("Enter file name for 1.1 file: "))
            inputfilename_13 = str(input("Enter file name for 1.3 file: "))
            return inputfilename_11, inputfilename_13
        elif user_input == "2":
            inputfilename_11 = "1.1 data Alva.xlsx"
            inputfilename_13 = "1.3 data Alva.xlsx"
            return inputfilename_11, inputfilename_13
        elif user_input == "0":
            sys.exit()
        else:
            print("Error! User command unrecognized.")

def Input_Sequence():
    while True:
        print("Please select a run sequence:")
        print("1) OFTpFpPpB")
        print("2) OFTpFpBPp")
        print("3) OFTpFpPp")
        print("4) OFTpFpB")
        print("0) Exit program")
        user_input = str(input("Enter input: "))
        if user_input == "1":
            return 1, "OFTpFpPpB"
        elif user_input == "2":
            return 2, "OFTpFpBPp"
        elif user_input == "3":
            return 3, "OFTpFpPp"
        elif user_input == "4":
            return 4, "OFTpFpB"
        elif user_input == "0":
            sys.exit()
        else:
            print("Error! User command unrecognized.")


# Name MatchWords
# The specialized matching function made for the 1.1/1.3 project.

def MatchWords(input_1, input_2):
    
    # Make sure any numbers turn into text.
    input_1 = str(input_1)
    input_2 = str(input_2)
    # Make filtered versions of inputs.
    input_3 = Filter(input_1, filterlist, lower=True)
    input_4 = Filter(input_2, filterlist, lower=True)
    
    try:
        if len(input_1) == len(input_2):
            tempwordlist = [input_1, input_2]
            # Take a word at random from the list:
            word_1 = tempwordlist.pop(choice(range(len(tempwordlist))))
            # Take the remaining word from the list:
            word_2 = tempwordlist.pop()
        elif len(input_1) > len(input_2):
            word_1 = input_2
            word_2 = input_1
        elif len(input_1) < len(input_2):
            word_1 = input_1
            word_2 = input_2
    except:
        print("String length measurement errors between", input_1, "and", input_2)
        mv = 0
        
    try:
        if len(input_3) == len(input_4):
            tempwordlist = [input_3, input_4]
            # Take a word at random from the list:
            word_3 = tempwordlist.pop(choice(range(len(tempwordlist))))
            # Take the remaining word from the list:
            word_4 = tempwordlist.pop()
        elif len(input_3) > len(input_4):
            word_3 = input_4
            word_4 = input_3
        elif len(input_3) < len(input_4):
            word_3 = input_3
            word_4 = input_4
    except:
        print("String length measurement errors between", input_3, "and", input_4)
        mv = 0
    
    len_1 = len(word_1)
    len_2 = len(word_2)
    len_3 = len(word_3)
    len_4 = len(word_4)
    
    # Uniltered match values:
    if word_1 in word_2:
        mv_unfil = len_1/len_2
    else:
        mv_unfil = max(SequenceMatcher(None, word_1, word_2).ratio(), 
                       SequenceMatcher(None, word_1.split(), 
                                       word_2.split()).ratio())
    
    # Filtered match values:
    if word_3 in word_4:
        mv_fil = len_3/len_4
    else:
        mv_fil = max(SequenceMatcher(None, word_3, word_4).ratio(), 
                     SequenceMatcher(None, word_3.split(), 
                                     word_4.split()).ratio())
    
    # Pick best match value:
    mv = max(mv_unfil, mv_fil)
        
    return mv

# Name: MatchBoxedLists
# A special matching function for the "box" format used in serial matching.
# Returns a "box" with matches.

def MatchBoxedLists(box, idx, orgcheck=False, outprint=False, randprint=False, 
                    filter_nullmatches=False, wordmatch=False, nummatch=False):
    
    best_intermediate_match_list = []   # List to store intermediate matches.
    match_value = 0                     # Current match value.
    best_match_value = 0                # Best match value so far.
    counter = 0
    current_best_match_list = []
    k = 0  # Progress counter. Used to print progress in percentage points.
    l = 0  # Box list counter. Used to iterate box lists that rows are compared to. 
    for row in box:
        # Unpack box
        row_1 = row[0]
        prev_match_value = row[1]
        list_2 = row[2]
        for row_2 in list_2:
            
            row2 = row_2
            
            # Org. Nr. special matching. Remove instances of "16" from beginning of 1.1 org.nrs.
            if orgcheck:
                if row_1[idx][0:2] == "16":  # If the first two numbers are 1 and 6:
                   row_1[idx] = row_1[idx][2:]  # Remove the first two digits.
                if row_2[idx][0:2] == "16":  # If the first two numbers are 1 and 6:
                   row_2[idx] = row_2[idx][2:]  # Remove the first two digits.
            
            if wordmatch:
                match_percentage = MatchWords(row_1[idx], row_2[idx])
            elif nummatch:
                try:
                    if float(row_1[idx]) > float(row_2[idx]):
                        longnum = float(row_1[idx])
                        shortnum = float(row_2[idx])
                    else:
                        longnum = float(row_2[idx])
                        shortnum = float(row_1[idx]) 
                    try:
                        match_percentage = shortnum/longnum
                    except:
                        match_percentage = 0
                except:
                    match_percentage = 0
            else:    
                if row_1[idx] == row_2[idx]:               
                    if randprint:
                        rand = choice(range(10))
                        if rand > 5:
                            counter += 1
                            print("Random match testing:", row_1[6], row_2[6])
                    match_percentage = 1
                else:
                    match_percentage = 0
                    
            if prev_match_value != 0:
                try:
                    match_value = round(match_percentage + prev_match_value, 4)/2
                    # if wordmatch:
                    #     match_value = match_percentage
                    # else:
                    #     if filter_nullmatches:
                    #         if match_percentage != 0:
                    #             match_value += round(match_percentage + prev_match_value, 4)/2 # * weight_vec[idx]
                    #         else:
                    #             match_value += match_percentage # * weight_vec[idx]
                    #     else:
                    #         match_value += round(match_percentage + prev_match_value, 4)/2 # * weight_vec[idx]
                except:
                    print("Hej")
                    match_value = 0
            else:
                match_value = match_percentage
            
            match_percentage = 0 
            
            # If current match value is equal to current best one, add to 
            # current list of best matches.
    
            if match_value == best_match_value and match_value != 0:
                    
                best_match_value = match_value
                current_best_match_list.append(row2)
    
                match_value = 0
                
            elif match_value > best_match_value: # If new value > best, also reset the intermediate lists.
                
                current_best_match_list = [] # reset best match list if new best match value > previous one.                
                
                best_match_value = match_value
                current_best_match_list.append(row2)
    
                match_value = 0    
                
            else:
    
                match_value = 0
    
        # Now add everything; match value, items from best_13 and best_11, into a single list.
        
        k += 1
        print("Progress:", round(k/len(box)*100, 2), "%")
        if filter_nullmatches:
            # If not 0 or better than or equal to previous match value - allow through.
            if best_match_value != 0 and best_match_value >= prev_match_value:
                best_intermediate_match_list.append([row[0], best_match_value, current_best_match_list])
        else:
            best_intermediate_match_list.append([row[0], best_match_value, current_best_match_list])
        best_match_value = 0  # Reset best match value for each new row_13 item.
        l += 1
        match_value = 0
        
    return best_intermediate_match_list

# Name: Read
# Basic read function. Reads excel files and returns sheet objects. 
# These need further processing to be useful inside Python.

def Read(filename):

    print("Reading!")

    workbook = load_workbook(filename=filename)
    sheet = workbook.active

    print("Reading complete!")

    return sheet

# Name: Unpack
# Unpacks items in multidimensional lists into a one-dimensional list.

def Unpack(packlist, index):
    unpacklist = []
    for i in range(len(packlist)):
        unpacklist.append(packlist[i][index])

    return unpacklist

# Name: WB_QuickPrint
# Prints all values within a workbook

def WB_QuickPrint(workbook, printnones):
    if printnones == None:
        printnones = True

    print("Printing...")

    for row in workbook.values:
       for value in row:
            if printnones:
                print(value)
            else:
                if value != None:
                    print(value)

# Name: WB_Print
# Prints all values within given range

def WB_Print(workbook, min_row, max_row, min_col, max_col, printnones):
    print("Printing...")
    for row in workbook.iter_rows(min_row=min_row, max_row=max_row, 
                                  min_col=min_col, max_col=max_col, 
                                  values_only=True):
       print(row)

# Name: Write
# Part of WriteTest_2 that actually does the writing to an excel sheet.

def Write(workbook, cellnames, values, filename):

    sheet = workbook.active

    for vals in range(len(values)):
        sheet[cellnames[vals]] = values[vals]

    workbook.save(filename=filename)
    print("Write successful!")
    
# ------------------- Section 4, main function code --------------------------
# The rest of the code is the "main" function code.
# -------------------- Subsection 4.1, processing ----------------------------
# This section reads in and processes the excel files, making them ready for
# matching.

# Run the user input function to determine file input

inputfilename_11, inputfilename_13 = Input_Name()

print("Using", inputfilename_11, "as 1.1 input.")
print("Using", inputfilename_13, "as 1.3 input.")

run_sequence, run_sequence_code = str(Input_Sequence())

print("Using", run_sequence_code, "as run sequence code.")

# Read the 1.1 and 1.3 excel databases and create corresponding workbook objects:

WB_11 = Read(inputfilename_11)
WB_13 = Read(inputfilename_13)

# Extract row and column amount:
r_max_1 = WB_11.max_row
r_max_2 = WB_13.max_row
c_max_1 = WB_11.max_column
c_max_2 = WB_13.max_column

# Extract the names of each column separately from the cell data.
# The output is a list with the columns names as separate items.

wb_11_names = Extract(WB_11, 1, 1, 1, c_max_1, False)
wb_13_names = Extract(WB_13, 1, 1, 1, c_max_2, False)

# Extract the rest of the rows and put them in a 2-dimensional list.

print("Extracting 1.1...")

for i in range(2, r_max_1+1):
    wb_11_item = Extract(WB_11, i, i, 1, c_max_1, False)  # Get 1.1 row
    for item in wb_11_item[9:12]:
        # To the end of the row, calculate and add the monetary sum:
        try:
            amount_sum += item
        except:
            pass  # If "None", do nothing.
    wb_11_item.append(amount_sum)  # Add sum to end of row.
    amount_sum = 0  # Reset sum counter.
    wb_11.append(wb_11_item)  # Add row to list of rows.

# Filter all 1.1 numbers that don't have an immediate connection to 1.3
# (every ID that doesn't have at least one "anslag" containing "ERUF")
# NOTE: This bit a not fully tested, but seems to work. Double check if time
# allows.
# NOTE 2: This function assumes that the excel sheet is sorted by "ärende-ID".
# NOTE 3: Excel has equivalent functions, or more "manual" filtering is possible in Excel using logic tables -
# a set of columns with true/false or 1/0 conditions.

if prefilter:
    print("Pre-filtering 1.1...")

    wb_11_filtered = []         # Create list which will contain filtered results.
    wb_11_copy = wb_11          # Copy to leave wb_11 intact
    filter_counter = 0
    wb_11_len = len(wb_11_copy)
    print(wb_11_len)
    tick = False                # Print first row that is popped.
    tock = False                # ERUF found at least once in a list of identical IDs
    tempfilterlist = []         # Temporary list used in the filtering loop

    while wb_11_copy:           # While the copy of wb_11 is not empty...
        tick = False            # Reset the state of "tick". Turn off to stop prints.
        tock = False            # Reset the state of "tock". Don't turn off.
        item = wb_11_copy.pop() # Pop a row from wb_11
        filter_counter += 1
        if tick:                # Optional single-time print. Use for debugging.
            tick = False
            print(item)
        if not prefilter_erufonly:
            for thing in wb_11_copy:        # Check entire wb_11.
                if item[0] == thing[0]:     # if same ID, append to temporary list.
                    tempfilterlist.append(thing)
        # Now, add these items to new filtered list and remove them from the
        # wb_11 copy:
            for i in tempfilterlist:     # Search filtered list.
                if "ERUF" in str(i[3]):  # if ERUF in any row, save the content of the temporary list.
                    for j in tempfilterlist:  # For every item in filter list
                        if not tock:
                            wb_11_filtered.append(j)  # Add items to final filtered list
                    tock = True
            # Remove item from wb_11_copy if found in temporary filter list.
                for k in wb_11_copy:
                    if i == k:
                        filter_index = wb_11_copy.index(k)
                        wb_11_copy.pop(filter_index)
                        filter_counter += 1
        else:
            if "ERUF" in str(item[3]):
                wb_11_filtered.append(item)
        tempfilterlist = [] # Reset filter list.
        print("Pre-filtering progress:",
        round(filter_counter/wb_11_len*100, 2), "%")
        #print(len(wb_11_copy))

    wb_11 = wb_11_filtered

print("Extracting 1.3...")

for i in range(2, r_max_2+1):
    wb_13_item = Extract(WB_13, i, i, 1, c_max_2, False)  # Get 1.3 row
    for item in wb_13_item[9:12]:
        # To the end of the row, calculate and add the monetary sum:
        try:
            amount_sum += item
        except:
            amount_sum = 0
            pass  # If "None", do nothing.
    wb_13_item.append(amount_sum)  # Add sum to end of row.
    amount_sum = 0  # Reset sum counter.
    wb_13.append(wb_13_item)

all_names.append("Match value")
for i in wb_11_names:  # Same here with 1.1 items.
    all_names.append(i)
all_names.append("Summa 1.1")  # Special case - a 1.1 sum value. 
for i in wb_13_names:  # Add all items in 1.3 to this new, longer list.
    all_names.append(i)
all_names.append("Summa 1.3")  # Special case - a 1.3 value.

# Now, the output structure is set in all_names and the lists to be matched are set in wb_11 and wb_13.

# --------------------- Subsection 4.2, matching -----------------------------
# This subsection actually performs the matching.

print("Matching...")

# This bit of code is useful for taking smaller parts of the 1.1 and 1.3 lists
# at random and comparing them instead. Useful for debugging.

# The number within range() is the number of tries the code will try to pick
# unique rows.

# wb_11_list = []
# for i in range(1000):
#     row = choice(wb_11)
#     if row not in wb_11_list:
#         wb_11_list.append(row)
# wb_13_list = []
# for i in range(1000):
#     row = choice(wb_13)
#     if row not in wb_13_list:
#         wb_13_list.append(row)

# Create initial input list:

# (Remember to use wb_11_list and wb_13_list instead if you want randomly
# picked rows from 1.1 and 1.3).

for row_11 in wb_11: 
    input_list.append([row_11, 0, wb_13])

# Mappings between the column names and their indices. Use these as the second
# argument for the MatchBoxedLists function to match with respect to that parameter:
# [0. ID], [1. Finansiär], [2. Finansiärstyp], [3. Anslag], [4. Beslut 1 år], 
# [5. Orgnamn], [6. Orgnr], [7. Projnamn], [8. Påg/Avsl.], 
# [9. EU(1.3)/Projmedel(1.1)], [10. Off.medel], [11. Priv.medel], [12. Summa].

# Some optional arguments for the MatchBoxedLists function:
# filter_nullmatches = True excludes best match values that equal zero.
# outprint = True means the matching progress will be shown as a percentage.
# orgcheck = True means the function will check if the input argument starts 
# with "16" and ignore it if that's the case. Use for orgnr only.
# wordmatch = True activates the partial string matching algorithm.
# nummatch = True activates matching by sums. mv = smaller value/larger value.

# If match = True in the global variable section, start matching:
if match:
    # Begin with the matching that is common to all run sequences:
    matchlist = MatchBoxedLists(input_list, 6, outprint=True, orgcheck=True,
                                wordmatch=False, filter_nullmatches=False)
    matchlist = MatchBoxedLists(matchlist, 2, wordmatch=True, outprint=True,
                                filter_nullmatches=False)
    matchlist = MatchBoxedLists(matchlist, 1, wordmatch=True, outprint=True,
                                filter_nullmatches=False)
    if run_sequence == "1":
        matchlist = MatchBoxedLists(matchlist, 7, wordmatch=True, outprint=True,
                                    filter_nullmatches=False)
        matchlist = MatchBoxedLists(matchlist, 12, nummatch=True, outprint=True)
    elif run_sequence == "2":
        matchlist = MatchBoxedLists(matchlist, 12, nummatch=True, outprint=True)
        matchlist = MatchBoxedLists(matchlist, 7, wordmatch=True, outprint=True,
                                    filter_nullmatches=False)
    elif run_sequence == "3":
        matchlist = MatchBoxedLists(matchlist, 7, wordmatch=True, outprint=True,
                                    filter_nullmatches=False)
    elif run_sequence == "4":
        matchlist = MatchBoxedLists(matchlist, 12, nummatch=True, outprint=True)
    else:
        print("Run sequence unrecognized. Skipping final parts of matching process.")
else:
    print("Matching turned off. Skipping matching.")

# Convert into output-friendly format:

output_list.append(all_names)

for thing in matchlist:
    for item in thing[2]:
        output_row.append(thing[1])
        for i in thing[0]:
            output_row.append(i)
        for i in item:
            output_row.append(i)
        output_list.append(output_row)
        output_row = []

# The postfilter is an attempt to only include unique matches. Works in theory, but could use additional testing in
# runs with different matching parameters.

if postfilter:
    counter = 0
    output_copy = output_list  # make a copy of the output list
    filtered_output_list = []
    filtered_output_list.append(output_list[0])  # add the names for the output list
    while output_copy and counter < (10*len(output_list)):
        item = output_copy.pop()
        if item not in filtered_output_list:
            filtered_output_list.append(item)
        counter += 1
    output_list = filtered_output_list
        


# --------------------- Subsection 4.3, printing -----------------------------
# This subsection prints the results from the matching.

# Write results in an excel sheet with the format (1.3 rows, match values,
# list of associated 1.1 items). Returns boolean that confirms that writing 
# was successful. Will attempt to make new file names until write is 
# successful or until a certain number of tries has been made.

while not successful and tries < 10:
    try:
        successful = ExcelWrite(1, 27, 1, len(output_list), output_list, outputfilename + ".xlsx")
        tries += 1
    except:
        filenamenumber += 1
        print("Filename already exists. Altering filename to", orig_filename + "_" + str(filenamenumber) + ".xlsx")
        outputfilename = orig_filename + "_" + str(filenamenumber)
        tries += 1
